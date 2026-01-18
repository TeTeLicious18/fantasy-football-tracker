import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import json
import os

def load_config():
    with open('config.json', 'r', encoding='utf-8') as f:
        return json.load(f)

def load_fixtures():
    try:
        with open('fixtures.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def create_excel():
    config = load_config()
    fixtures = load_fixtures()
    players = config['players']
    league_name = config['league']['name']
    start_md = config['matchdays']['start']
    end_md = config['matchdays']['end']
    
    # Create workbook
    if os.path.exists('predictions.xlsx'):
        wb = openpyxl.load_workbook('predictions.xlsx')
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
    
    # Styles
    title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    header_font = Font(bold=True)
    match_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    score_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    points_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    total_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')
    
    sheet_names = []
    
    # Create matchday sheets
    for md in range(start_md, end_md + 1):
        md_str = str(md)
        sheet_name = f"Matchday {md}"
        sheet_names.append(sheet_name)
        
        # Remove if exists
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        
        ws = wb.create_sheet(sheet_name)
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"MATCHDAY {md} - {league_name.upper()}"
        ws['A1'].fill = title_fill
        ws['A1'].font = title_font
        ws['A1'].alignment = center
        
        row = 3
        
        # Get matches for this matchday
        matches = fixtures.get(md_str, [])
        
        if not matches:
            # Create 8 placeholder matches
            matches = [{"home": f"Team {i*2+1}", "away": f"Team {i*2+2}", "date": ""} for i in range(8)]
        
        for match in matches:
            home = match.get('home', 'Home Team')
            away = match.get('away', 'Away Team')
            match_date = match.get('date', '')
            
            # Date/Time row
            ws[f'A{row}'] = "Date/Time:"
            ws[f'A{row}'].font = Font(italic=True, size=10)
            ws.merge_cells(f'B{row}:C{row}')
            ws[f'B{row}'] = match_date
            ws[f'B{row}'].border = border
            ws[f'B{row}'].alignment = center
            ws[f'B{row}'].font = Font(bold=True, size=10)
            row += 1
            
            # Match header
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = f"{home} vs {away}"
            ws[f'A{row}'].fill = match_fill
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws[f'A{row}'].border = border
            
            ws[f'D{row}'] = "Real Score:"
            ws[f'D{row}'].font = header_font
            ws[f'D{row}'].alignment = center
            
            ws[f'E{row}'].fill = score_fill
            ws[f'E{row}'].border = border
            ws[f'F{row}'].fill = score_fill
            ws[f'F{row}'].border = border
            
            real_home = f'E{row}'
            real_away = f'F{row}'
            
            row += 2
            
            # Headers
            for col, hdr in enumerate(['Player', 'Pred Home', 'Pred Away', 'Points'], 1):
                cell = ws.cell(row=row, column=col)
                cell.value = hdr
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border
            row += 1
            
            # Player rows
            for player in players:
                ws.cell(row=row, column=1).value = player
                ws.cell(row=row, column=1).border = border
                ws.cell(row=row, column=1).alignment = center
                
                ws.cell(row=row, column=2).border = border
                ws.cell(row=row, column=2).alignment = center
                
                ws.cell(row=row, column=3).border = border
                ws.cell(row=row, column=3).alignment = center
                
                # Points formula
                ph = f'B{row}'
                pa = f'C{row}'
                
                exact = f"AND({real_home}={ph},{real_away}={pa})"
                home_win = f"AND({real_home}>{real_away},{ph}>{pa})"
                away_win = f"AND({real_home}<{real_away},{ph}<{pa})"
                draw = f"AND({real_home}={real_away},{ph}={pa})"
                
                formula = f'=IF(OR(ISBLANK({real_home}),ISBLANK({real_away}),ISBLANK({ph}),ISBLANK({pa})),"",IF({exact},3,IF(OR({home_win},{away_win},{draw}),1,0)))'
                
                ws.cell(row=row, column=4).value = formula
                ws.cell(row=row, column=4).fill = points_fill
                ws.cell(row=row, column=4).border = border
                ws.cell(row=row, column=4).alignment = center
                
                row += 1
            
            row += 2
        
        # Sheet total
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = f"TOTAL MATCHDAY {md}"
        ws[f'A{row}'].fill = total_fill
        ws[f'A{row}'].font = total_font
        ws[f'A{row}'].alignment = center
        
        row += 2
        ws[f'A{row}'] = "Player"
        ws[f'B{row}'] = "Points"
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].fill = header_fill
        ws[f'B{row}'].font = header_font
        
        row += 1
        for player in players:
            ws[f'A{row}'] = player
            ws[f'A{row}'].border = border
            ws[f'B{row}'] = f'=SUMIF(A:A,"{player}",D:D)'
            ws[f'B{row}'].border = border
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
        
        print(f"  Created {sheet_name}")
    
    # Create LEADERBOARD sheet
    if "LEADERBOARD" in wb.sheetnames:
        del wb["LEADERBOARD"]
    
    summary = wb.create_sheet("LEADERBOARD")
    
    summary.column_dimensions['A'].width = 15
    for i in range(1, len(sheet_names) + 3):
        summary.column_dimensions[get_column_letter(i+1)].width = 12
    
    summary.merge_cells(f'A1:{get_column_letter(len(sheet_names)+2)}1')
    summary['A1'] = f"LEADERBOARD - {league_name.upper()}"
    summary['A1'].fill = title_fill
    summary['A1'].font = title_font
    summary['A1'].alignment = center
    
    # Headers
    summary['A3'] = "Player"
    summary['A3'].fill = header_fill
    summary['A3'].font = header_font
    
    for i, sn in enumerate(sheet_names):
        cell = summary.cell(row=3, column=i+2)
        cell.value = f"MD{start_md + i}"
        cell.fill = header_fill
        cell.font = Font(bold=True, size=9)
        cell.alignment = center
    
    total_col = len(sheet_names) + 2
    summary.cell(row=3, column=total_col).value = "TOTAL"
    summary.cell(row=3, column=total_col).fill = total_fill
    summary.cell(row=3, column=total_col).font = total_font
    
    # Player rows
    for p_idx, player in enumerate(players):
        row = 4 + p_idx
        summary.cell(row=row, column=1).value = player
        summary.cell(row=row, column=1).border = border
        
        for s_idx, sn in enumerate(sheet_names):
            formula = f"=SUMIF('{sn}'!A:A,\"{player}\",'{sn}'!D:D)"
            cell = summary.cell(row=row, column=s_idx+2)
            cell.value = formula
            cell.border = border
            cell.alignment = center
        
        # Total
        start_col = get_column_letter(2)
        end_col = get_column_letter(len(sheet_names)+1)
        total_formula = f"=SUM({start_col}{row}:{end_col}{row})"
        total_cell = summary.cell(row=row, column=total_col)
        total_cell.value = total_formula
        total_cell.font = Font(bold=True, size=12)
        total_cell.border = border
        total_cell.alignment = center
    
    print("  Created LEADERBOARD")
    
    wb.save('predictions.xlsx')
    print(f"\nExcel 'predictions.xlsx' created with {len(sheet_names)} matchday sheets!")

if __name__ == "__main__":
    print("=" * 50)
    print("PREDICTIONS EXCEL GENERATOR")
    print("=" * 50)
    create_excel()
    input("\nPress Enter to close...")
