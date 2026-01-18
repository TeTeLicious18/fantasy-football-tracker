import openpyxl
import requests
from bs4 import BeautifulSoup
import json
import re
import time
import sys

sys.stdout.reconfigure(encoding='utf-8')

def load_config():
    with open('config.json', 'r', encoding='utf-8') as f:
        return json.load(f)

def get_all_scores(league_id):
    """Fetch scores from livescore.bz for a specific league"""
    print("Fetching scores from livescore.bz...")
    
    LEAGUE_IDS = {
        "romania/superliga": 131,
        "england/premier-league": 2,
        "spain/la-liga": 8,
        "germany/bundesliga": 9,
        "italy/serie-a": 10,
        "france/ligue-1": 11,
    }
    
    lid = LEAGUE_IDS.get(league_id.lower(), 131)
    
    urls = [
        "https://www.livescore.bz/",
        f"https://www.livescore.bz/en/football/league/{lid}/results/"
    ]
    
    all_matches = []
    headers = {"User-Agent": "Mozilla/5.0"}
    
    for url in urls:
        try:
            resp = requests.get(url, headers=headers, timeout=10)
            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, 'html.parser')
                for a in soup.find_all('a', href=True):
                    if 'event' in a['href']:
                        text = a.get_text(" ", strip=True)
                        all_matches.append(text)
            time.sleep(0.5)
        except Exception as e:
            print(f"Error: {e}")
    
    print(f"Found {len(all_matches)} match entries")
    return all_matches

def parse_score(text):
    match = re.search(r'(\d+)\s*-\s*(\d+)', text)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None, None

def check_team(team_name, text):
    text_lower = text.lower()
    if team_name.lower()[:4] in text_lower:
        return True
    return False

def find_score(home, away, web_data):
    for line in web_data:
        if check_team(home, line) and check_team(away, line):
            return parse_score(line)
    return None, None

def update_all_sheets():
    config = load_config()
    league_id = config['league']['flashscore_id']
    
    file_path = 'predictions.xlsx'
    
    try:
        open(file_path, 'a+').close()
    except PermissionError:
        print("[ERROR] Excel file is open! Close it and try again.")
        input("Press Enter to exit...")
        return
    
    web_data = get_all_scores(league_id)
    
    try:
        wb = openpyxl.load_workbook(file_path)
        total_updates = 0
        
        print("\nUpdating scores in all sheets...\n")
        
        for sheet_name in wb.sheetnames:
            if "LEADERBOARD" in sheet_name.upper():
                continue
            
            sheet = wb[sheet_name]
            sheet_updates = 0
            
            print(f"--- {sheet_name} ---")
            
            for row in sheet.iter_rows(min_row=1, max_col=6, values_only=False):
                if row[3] and row[3].value == "Real Score:":
                    title = row[0].value
                    if title and " vs " in str(title):
                        h, a = str(title).split(" vs ")
                        sh, sa = find_score(h.strip(), a.strip(), web_data)
                        
                        if sh is not None:
                            print(f"  [+] {title}: {sh}-{sa}")
                            row[4].value = sh
                            row[5].value = sa
                            sheet_updates += 1
                            total_updates += 1
                        else:
                            if row[4].value is not None:
                                print(f"  [=] {title}: keeping ({row[4].value}-{row[5].value})")
                            else:
                                print(f"  [-] {title}: not found")
            
            if sheet_updates > 0:
                print(f"  Updates: {sheet_updates}")
        
        wb.save(file_path)
        print(f"\n{'='*50}")
        print(f"TOTAL UPDATES: {total_updates}")
        print(f"{'='*50}")
    except Exception as e:
        print(f"Error: {e}")
    
    input("\nPress Enter to close...")

if __name__ == "__main__":
    print("=" * 50)
    print("SCORE UPDATER")
    print("=" * 50)
    update_all_sheets()
